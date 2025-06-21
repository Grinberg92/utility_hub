import DaVinciResolveScript as dvr
import re
from threading import Thread
import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dvr_tools.logger_config import get_logger
import pandas as pd
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import Qt
from threading import Thread
import sys
from dvr_tools.css_style import apply_style

logger = get_logger(__file__)

class GUI(QtWidgets.QWidget):

    finished_signal = QtCore.pyqtSignal(str, str)
    error_signal = QtCore.pyqtSignal(str, str)

    def __init__(self):
        super().__init__()
        self.setWindowTitle('Color&FPS Setter')
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.resize(450, 130)
        self.clip_color_list = ['Orange', 'Yellow', 'Lime', 'Teal', 'Green', 'Purple', 'Navy',
                                'Apricot', 'Olive', 'Violet', 'Blue', 'Pink', 'Tan', 'Beige',
                                'Brown', 'Chocolate']

        self.init_ui()
        self.finished_signal.connect(self.show_message_box)
        self.error_signal.connect(self.show_message_box)

    def init_ui(self):
        layout = QtWidgets.QVBoxLayout()

        # === Ряд с чекбоксами и FPS-инпутом ===
        options_layout = QtWidgets.QHBoxLayout()

        self.checkbox_color = QtWidgets.QCheckBox("Set color")
        self.checkbox_fps = QtWidgets.QCheckBox("Set FPS")
        self.checkbox_excel = QtWidgets.QCheckBox("Create Excel")

        # FPS label + input в одном layout
        fps_layout = QtWidgets.QHBoxLayout()
        fps_label = QtWidgets.QLabel("FPS:")
        fps_label.setContentsMargins(0, 0, 2, 0)  # Минимальный отступ
        self.fps_entry = QtWidgets.QLineEdit()
        self.fps_entry.setFixedWidth(40)
        self.fps_entry.setText("24")

        fps_layout.addWidget(fps_label)
        fps_layout.addWidget(self.fps_entry)
        fps_widget = QtWidgets.QWidget()
        fps_widget.setLayout(fps_layout)

        # Применяем расширяющиеся size policy
        for widget in [self.checkbox_color, self.checkbox_fps, self.checkbox_excel, fps_widget]:
            widget.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)

        # Добавляем всё в layout
        options_layout.addWidget(self.checkbox_color)
        options_layout.addWidget(self.checkbox_fps)
        options_layout.addWidget(self.checkbox_excel)
        options_layout.addWidget(fps_widget)

        self.checkbox_color.stateChanged.connect(self.update_input_state)
        self.checkbox_excel.stateChanged.connect(self.update_input_state)
        self.checkbox_fps.stateChanged.connect(self.update_input_state)

        # === Ряд с шириной и высотой ===
        resolution_layout = QtWidgets.QHBoxLayout()
        resolution_layout.setContentsMargins(0, 0, 0, 0)
        resolution_layout.setSpacing(5)  # Уменьшаем отступ между виджетами

        self.width_entry = QtWidgets.QLineEdit("2048")
        self.width_entry.setPlaceholderText("Widthа")
        self.width_entry.setMaximumWidth(80)  # или setFixedWidth

        self.separator = QtWidgets.QLabel("x")

        self.height_entry = QtWidgets.QLineEdit("858")
        self.height_entry.setPlaceholderText("Height")
        self.height_entry.setMaximumWidth(80)

        resolution_layout.addWidget(QtWidgets.QLabel("Resolution:"))
        resolution_layout.addWidget(self.width_entry)
        resolution_layout.addWidget(self.separator)
        resolution_layout.addWidget(self.height_entry)

        # Оборачиваем в виджет, чтобы выравнивать
        resolution_widget = QtWidgets.QWidget()
        resolution_widget.setLayout(resolution_layout)

        # --- Путь рендера ---
        path_layout = QtWidgets.QHBoxLayout()
        self.path_input = QtWidgets.QLineEdit()
        self.browse_btn = QtWidgets.QPushButton("Choose")
        self.browse_btn.clicked.connect(self.select_file)
        path_layout.addWidget(QtWidgets.QLabel("Excel Path:"))
        path_layout.addWidget(self.path_input)
        path_layout.addWidget(self.browse_btn)

        # === Кнопка запуска ===
        self.run_button = QtWidgets.QPushButton("Start")
        self.run_button.clicked.connect(self.on_run_clicked)

        # === Добавление в layout ===
        layout.addLayout(options_layout)
        layout.addWidget(resolution_widget, alignment=QtCore.Qt.AlignCenter)
        layout.addLayout(path_layout)
        layout.addWidget(self.run_button)

        self.setLayout(layout)
        self.update_input_state()  # Инициализируем доступность FPS поля

    def select_file(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Выбор файла Excel",
            filter="Excel Files (*.xlsx);;All Files (*)",
            directory="Exel_Resolution_Spreadsheet.xlsx"  # имя по умолчанию
        )
        if path:
            # Убедимся, что расширение есть
            if not path.endswith(".xlsx"):
                path += ".xlsx"
            self.path_input.setText(path)

    def show_message_box(self, title, message):
        if title == "Ошибка":
            QMessageBox.critical(self, title, message)
        else:
            QMessageBox.information(self, title, message)

    def update_input_state(self):

        # Блокировка инпута FPS
        if not self.checkbox_fps.isChecked():
            self.fps_entry.setEnabled(False)
        else:
            self.fps_entry.setEnabled(True)

        # Блокировка инпута Resolution
        if not self.checkbox_excel.isChecked():
            self.width_entry.setEnabled(False)
            self.height_entry.setEnabled(False)
            self.path_input.setEnabled(False)
            self.browse_btn.setEnabled(False)
        else:
            self.width_entry.setEnabled(True)
            self.height_entry.setEnabled(True)
            self.path_input.setEnabled(True)
            self.browse_btn.setEnabled(True)

    def on_run_clicked(self):

        logger.debug("Запуск скрипта")
        
        self.run_button.setEnabled(False)

        # Получаем флаги с чекбоксов
        self.run_coloring = self.checkbox_color.isChecked()
        self.create_excel = self.checkbox_excel.isChecked()
        self.set_fps = self.checkbox_fps.isChecked()
        self.fps_input = self.fps_entry.text()
        self.exel_folder = self.path_input.text()
        self.output_res_height = self.height_entry.text()
        self.output_res_width = self.width_entry.text()

        if not self.exel_folder and self.create_excel:
            QMessageBox.warning(self, "Предупреждение", "Пожалуйста, укажите путь для таблицы")
            logger.warning("Пожалуйста, укажите путь для таблицы")
            self.run_button.setEnabled(True)
            return 

        logger.debug("\n".join((f"SetUp:", f"Set color: {self.run_coloring}", f"Set FPS: {self.set_fps}",
                               f"FPS: {self.fps_entry}", f"Create Excel: {self.create_excel}",
                               f"Resolution: {self.output_res_width}x{self.output_res_height}",
                               f"Excel Path: {self.exel_folder}")))
        
        Thread(target=self.run_script_wrapper).start()

    def run_script_wrapper(self):
        try:
            fps_value = self.fps_entry.text() if self.set_fps else None
            self.run_da_vinci_script(fps_value, create_exel=self.create_excel, run_coloring=self.run_coloring)
        finally:
            self.run_button.setEnabled(True)

    def run_da_vinci_script(self, fps_value, create_exel, run_coloring=True):
        """Основная логика скрипта для DaVinci Resolve"""

        def get_spreadsheet_data(data):
            '''
            Функция вычисляет/получает данные для экспорта в EXEL таблицу 
            '''
            def export_to_exel(table_data_list):
                '''
                Функция экспортирует данные в EXEL таблицу
                '''
                COLOR_HEX_MAP = {
                    'Orange': "FFA500",
                    'Yellow': "FFFF99",
                    'Lime': "BFFF00",
                    'Teal': "008080",
                    'Green': "66CC66",
                    'Purple': "9370DB",
                    'Navy': "000080",
                    'Apricot': "FBCEB1",
                    'Olive': "808000",
                    'Violet': "EE82EE",
                    'Blue': "87CEEB",
                    'Pink': "FFC0CB",
                    'Tan': "D2B48C",
                    'Beige': "F5F5DC",
                    'Brown': "A52A2A",
                    'Chocolate': "D2691E"
                }

                def style_excel_table(ws, table_start_row, num_columns):
                    fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    for row in ws.iter_rows(min_row=table_start_row, max_row=ws.max_row, min_col=1, max_col=num_columns):
                        for cell in row:
                            cell.fill = fill
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    for cell in ws[table_start_row]:
                        cell.font = Font(bold=True)

                    for col_idx in range(1, num_columns + 1):
                        col_letter = get_column_letter(col_idx)
                        max_len = max(
                            len(str(ws.cell(row=row, column=col_idx).value)) for row in range(table_start_row, ws.max_row + 1)
                        )
                        ws.column_dimensions[col_letter].width = max(12, max_len + 2)

                def color_rows_by_color_column(ws, color_column_name="Цвет", start_row=4):
                    color_col_idx = None
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=3, column=col).value == color_column_name:
                            color_col_idx = col
                            break

                    if color_col_idx is None:
                        print("Колонка 'Цвет' не найдена.")
                        return

                    for row in range(start_row, ws.max_row + 1):
                        color_name = ws.cell(row=row, column=color_col_idx).value
                        hex_color = COLOR_HEX_MAP.get(color_name)
                        if hex_color:
                            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                            cell = ws.cell(row=row, column=color_col_idx)
                            cell.fill = fill

                headers = ["Цвет", "Камера", "Оптика", "Исходное разрешение",
                            "Разрешение выдачи", "Разрешение 1.5x", "Разрешение 2x",
                            "Mxf для AVID", "Доп информация"]
                table_data_list.insert(0, headers)

                expected_columns = len(headers)
                for row in table_data_list[1:]:
                    while len(row) < expected_columns:
                        row.append("")

                df = pd.DataFrame(table_data_list[1:], columns=table_data_list[0])
                df.to_excel(self.exel_folder, index=False, startrow=2)

                wb = openpyxl.load_workbook(self.exel_folder)
                ws = wb.active

                num_columns = len(headers)

                ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_columns)
                header_cell = ws.cell(row=1, column=1)
                header_cell.value = "PROJECT NAME HERE"
                header_cell.font = Font(bold=True, size=14)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")

                style_excel_table(ws, table_start_row=3, num_columns=num_columns)

                color_rows_by_color_column(ws, color_column_name="Цвет", start_row=4)

                # >>> ОБЪЕДИНЕНИЕ колонки "Доп информация"
                for col_idx in range(1, num_columns + 1):
                    if ws.cell(row=3, column=col_idx).value == "Доп информация":
                        info_col = col_idx
                        break
                else:
                    info_col = None

                if info_col:
                    start_row = 4
                    end_row = ws.max_row
                    ws.merge_cells(start_row=start_row, start_column=info_col, end_row=end_row, end_column=info_col)

                    merged_cell = ws.cell(row=start_row, column=info_col)
                    merged_cell.value = ""  # или можно вставить любую строку
                    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
                    merged_cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                    merged_cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                wb.save(self.exel_folder)
                logger.info(f"Таблица Excel успешно сформировнаю Путь: {self.exel_folder}")

            def rescale_resolution(width, height, aspect):
                '''
                Функция получает разрешение для стандартной 2к выдачи
                '''
                if aspect == "Square":
                    calculate_height = int((math.ceil((height * target_resolution_width / width) / 2) * 2))
                    return (target_resolution_width, calculate_height)
                else:
                    calculate_width = int((math.ceil((width * target_resolution_height / height) / 2) * 2))
                    return (calculate_width, target_resolution_height)
                
            def rescale_1_5_and_2_x(width, height, mult=None):
                '''
                Функция пересчитывает стандартное 2к разрешение в 1.5 кратноме и 2 кратное разрешение
                '''
                result_width = int((math.ceil((width * mult) / 2)) * 2)
                result_height = int((math.ceil((height * mult) / 2)) * 2)
                return result_width, result_height
            
            def rescale_avid(width, height):

                avid_target_resolution_width = 1920
                avid_target_resolution_height = 1080

                calculate_height = int((math.ceil((height * avid_target_resolution_width / width) / 2) * 2))
                return (avid_target_resolution_width, calculate_height)
                
            
            # Целевое финальное разрешение
            target_resolution_height = int(self.output_res_height)
            target_resolution_width = int(self.output_res_width)

            # Получили сортировку и цвета такие же как для расцветовки OCF
            data_zip = list(zip(self.clip_color_list, sorted(data.items(), key=lambda x: len(x[1]), reverse=True)))

            # Сбор всех возможных разрешений и данных для экспорта в таблицу
            table_data_list = []
            for color, (resolution_aspect, clips) in data_zip:
                resolution, aspect = resolution_aspect
                camera_resolution_width, camera_resolution_heigth = list(map(int, resolution.split("x")))
                first_latter_clip_name = list(set([clip.GetName()[0].upper() for clip in clips]))
                avid_scale_resolution_width, avid_scale_resolution_height = rescale_avid(camera_resolution_width, camera_resolution_heigth)
                scale_2k_resolution_width, scale_2k_resolution_height = rescale_resolution(camera_resolution_width, camera_resolution_heigth, aspect)
                scale_1_5_resolution_width, scale_1_5_resolution_height = rescale_1_5_and_2_x(scale_2k_resolution_width, scale_2k_resolution_height, mult=1.5)
                scale_2_resolution_width, scale_2_resolution_height = rescale_1_5_and_2_x(scale_2k_resolution_width, scale_2k_resolution_height, mult=2)
                if aspect == "Square":
                    aspect = "Сферическая"
                else:
                    aspect = "Анаморф"
                table_data_list.append([color,
                      first_latter_clip_name, 
                      aspect,
                      f"{camera_resolution_width}x{camera_resolution_heigth}", 
                      f"{scale_2k_resolution_width}x{scale_2k_resolution_height}",
                      f"{scale_1_5_resolution_width}x{scale_1_5_resolution_height}", 
                      f"{scale_2_resolution_width}x{scale_2_resolution_height}",
                      f"{avid_scale_resolution_width}x{avid_scale_resolution_height}"
                      ])
            export_to_exel(table_data_list)

        try:
            resolve = dvr.scriptapp("Resolve")
            project = resolve.GetProjectManager().GetCurrentProject()
            media_pool = project.GetMediaPool()

            target_bin = self.find_target_bin(media_pool)  # Ищем папку OCF
            if not target_bin:
                self.finished_signal.emit("Ошибка", "Папка OCF не найдена.")
                logger.critical('Папка OCF не найдена.')
                return

            clips_dict = {}  # Словарь с данными разрешение: список клипов с таким разрешением
            spreadsheet_info_dict = {} # Словарь с данными (разрешение, аспект): список клипов с таким разрешением
            clips = self.get_clips_from_bin(target_bin)

            for clip in clips:
                if clip.GetName() != '' and clip.GetName().lower().endswith((".mxf", ".braw", ".arri", ".mov", ".r3d", ".mp4", ".dng", ".jpg", ".cine")):
                    # Находит анаморф, вычисляет ширину по аспекту
                    if clip.GetClipProperty('PAR') != 'Square' and clip.GetClipProperty('PAR'):
                        # Меняем FPS если не соответствует проектному и не выбрано создание таблицы
                        if float(clip.GetClipProperty("FPS")) != float(fps_value) and self.set_fps:
                            clip.SetClipProperty("FPS", "24")
                            logger.debug(f"Изменен FPS на {fps_value} для клипа {clip.GetName()}")

                        aspect = clip.GetClipProperty('PAR')
                        width, height = clip.GetClipProperty('Resolution').split('x')
                        resolution = "x".join([str(width), str(int((int(height) / float(aspect)) + (int(height) / float(aspect)) % 2))])
                        clips_dict.setdefault(resolution, []).append(clip)
                        spreadsheet_info_dict.setdefault((resolution, aspect), []).append(clip) # Данные для таблицы
                    else:
                        # Меняем FPS если не соответствует проектному и не выбрано создание таблицы
                        if float(clip.GetClipProperty("FPS")) != float(fps_value) and self.set_fps:
                            clip.SetClipProperty("FPS", "24")
                            logger.debug(f"Изменен FPS на {fps_value} для клипа {clip.GetName()}")

                        aspect = clip.GetClipProperty('PAR')
                        clips_dict.setdefault(clip.GetClipProperty('Resolution'), []).append(clip)
                        spreadsheet_info_dict.setdefault((clip.GetClipProperty('Resolution'), aspect), []).append(clip) # Данные для таблицы

            # Опционально получаем данные для таблицы
            if create_exel:
                get_spreadsheet_data(spreadsheet_info_dict)

            # Опционально запускаем присваивание цвета клипам в медиапуле и устанавливаем проектный FPS
            if run_coloring:
                # Сортировка по количеству клипов в группе разрешения
                sorted_clip_list = list(zip(self.clip_color_list, {res: clips for res, clips in sorted(clips_dict.items(), key=lambda x: len(x[1]), reverse=True) if res != ""}))

                # Установка отдельного цвета на каждую группу разрешений
                for color, res in sorted_clip_list:
                    if res in clips_dict:
                        for clip in clips_dict[res]:
                            clip.SetClipColor(color)
                        logger.debug(f"Установлен цвет {color} на группу разрешения {res}")

            self.finished_signal.emit("Успех", f"Обработка закончена.")
            logger.debug(f"Обработка закончена.")
            
        except Exception as e:
            self.error_signal.emit("Ошибка", f"Произошла ошибка: {str(e)}")
            logger.exception(f"Произошла ошибка: {str(e)}")

    def get_clips_from_bin(self, bin):
        """Рекурсивно получает все клипы из папки и её подпапок."""
        clips = list(bin.GetClipList())  # Получаем клипы из текущей папки
        for sub_bin in bin.GetSubFolderList():  # Рекурсивно проходим по подпапкам
            clips.extend(self.get_clips_from_bin(sub_bin))
        return clips

    def find_target_bin(self, media_pool, target_name="OCF"):
        """Ищет папку с именем target_name в Media Pool."""
        root_bin = media_pool.GetRootFolder()
        return self.search_bin_recursive(root_bin, target_name)

    def search_bin_recursive(self, bin, target_name):
        """Рекурсивный поиск папки в Media Pool."""
        if re.search(target_name.lower(), bin.GetName().lower()):
            return bin
        for sub_bin in bin.GetSubFolderList():
            result = self.search_bin_recursive(sub_bin, target_name)
            if result:
                return result
        return None

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    apply_style(app)
    window = GUI()
    window.show()
    sys.exit(app.exec_())



