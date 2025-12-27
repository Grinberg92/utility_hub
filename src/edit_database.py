from timecode import Timecode as tc
import re
import sys
import os
import json
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime as dt, date as d
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject, QSize, QUrl
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import (QMessageBox, QVBoxLayout, QHBoxLayout, QLabel, 
    QLineEdit, QPushButton,  QApplication, QFileDialog, QWidget, QTextEdit, 
    QTabWidget, QComboBox,  QTreeView, QHeaderView, QTextBrowser)
from timecode import Timecode as tc
from dvr_tools.css_style import apply_style
from common_tools.edl_parsers import detect_edl_parser, EDLParser
from dvr_tools.logger_config import get_logger
from config.config_loader import load_config
from config.config import get_config
from config.global_config import GLOBAL_CONFIG

logger = get_logger(__file__)

DATA_PATH = {"win32": GLOBAL_CONFIG["paths"]["editdatabase_path_win"], 
                        "darwin": GLOBAL_CONFIG["paths"]["editdatabase_path_mac"]}[sys.platform]

def get_output_path(project: str, ext: str, report_name: str, subfolder=None) -> str:
    """
    Получение пути к бекапу отчета проверки секвенций.
    """
    date = dt.now().strftime("%Y%m%d")

    output_path = (
        Path(
            {"win32": GLOBAL_CONFIG["paths"]["root_projects_win"],
            "darwin": GLOBAL_CONFIG["paths"]["root_projects_mac"]}[sys.platform]
        )
        / project
        / GLOBAL_CONFIG["output_folders"]["edit_database"] / date / (subfolder if subfolder is not None else "")
        / f"{report_name}_{date}.{ext}"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path

class EditDatabase:
    """
    Класс базы данных монтажей.
    """
    def __init__(self, data_base_path: str, project: str):
        self.data_base = data_base_path
        self.project = project

        if os.path.exists(data_base_path):
            with open(data_base_path, "r", encoding="utf-8") as f:
                content = f.read().strip()
                if not content:
                    self.data = {} 
                else:
                    self.data = json.loads(content)
        else:
            self.data = {}

        if project not in self.data:
            self.data[project] = {}

    def add_shot(self, project: str, shot_name: str, edit_name: str, shot_id: str,
                track_type: str, transition: str, 
                src_in: str, src_out: str, src_out_full: str, rec_in: str, 
                rec_out: str, src_name: str, update_status) -> None:
        """
        Добавляет новый монтаж шота.
        Проверяет существует ли проект или нет.

        :param update_status: Булево значение, указывающее добавлять ли текущий монтаж в статус 'actual' или нет.
        """
        
        if shot_name not in self.data[project]:
            self.data[project][shot_name] = {}

        if update_status:
            for edit in self.data[project][shot_name].values():
                edit["is_actual"] = False

        self.data[project][shot_name][edit_name] = {
            "id": shot_id,
            "shot_name": shot_name,
            "src_name": src_name,
            "track_type": track_type,
            "transition": transition,
            "src_in": src_in,
            "src_out": src_out,
            "src_out_full": src_out_full,
            "rec_in": rec_in,
            "rec_out": rec_out,
            "is_actual": update_status,
            "edit_version": edit_name,
            "add_data": dt.strftime(dt.today(), "%Y-%m-%d %H:%M:%S")
        }

    def _remove_edit(self, project: str, input_edit) -> None:
        """
        Удаляет все вхождения монтажа input_edit в указанном проекте.
        """
        if project not in self.data:
            return False

        project_data = self.data[project]

        shots_to_clean = []
        for shot, edits in project_data.items():

            if input_edit in edits:
                del edits[input_edit]

            if not edits:
                shots_to_clean.append(shot)

        for shot in shots_to_clean:
            del project_data[shot]

        self.save()

    def _remove_project(self, project: str) -> bool | None:
        """
        Удаляет проект.
        """
        if project not in self.data:
            return False
        
        del self.data[project]

        self.save()

    def _remove_shots(self, project: str, shots: list) -> bool | None:
        """
        Удаляет шоты.
        """
        if project not in self.data:
            return False
        shots = shots.split(" ")
        for shot in shots:
            if shot in self.data[project]:
                del self.data[project][shot]

        self.save()

    def get_shots_by_edit(self, project: str, input_edit_name: str) -> dict:
        """
        Возвращает словарь с шотами отфильтрованными по названию входящего монтажа.
        """
        result_data = {}
        project_data = self.data.get(project, {})
        for shot, edit in project_data.items():
            for edit_name, edit_data in edit.items():
                if edit_name == input_edit_name:
                    result_data.update({shot: edit_data})
        
        return result_data
    
    def get_shots_by_actual(self, project: str) -> dict:
        """
        Возвращает шоты со статусом монтажа "is_actual": True.
        """
        result_data = {}
        project_data = self.data.get(project, {})
        for shot, edit in project_data.items():
            for _, edit_data in edit.items():
                if edit_data.get("is_actual"):
                    result_data.update({shot: edit_data})
        
        return result_data
    
    def get_shots_by_edits(self, project: str, input_edits: list) -> dict:
        """
        Возвращает словарь с шотами отфильтрованными по названию входящих монтажей.
        Значения ключей содержат значения всех монтажей в параметре input_edits.
        """
        result_data = {}
        project_data = self.data.get(project, {})
        for input_edit in input_edits:
            for shot, edit in project_data.items():
                for edit_name, edit_data in edit.items():
                    if edit_name == input_edit:
                        result_data.setdefault(shot, []).append(edit_data)
        
        return result_data
    
    def get_edits(self, project: str) -> list:
        """
        Возвращает список всех монтажей добавленных в базу данных.
        """
        names = Counter()
        shots = list(self.data[project].values())
        for shot_edit in shots:
            for edit_name, _ in shot_edit.items():
                names.update([edit_name])
        return list(names)

    def save(self):
        """
        Сохранить базу данных.
        """
        with open(self.data_base, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=4, ensure_ascii=False)

    def backup(self):
        """
        Бэкап базы данных.
        """
        backup_path = self.data_base.replace('.json', '_backup.json')
        with open(backup_path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=4, ensure_ascii=False)       

    def _clear(self):
        """
        Очистить базу данных.
        """
        self.data = {}
        with open(self.data_base, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=4, ensure_ascii=False) 

class ShotRestorer(QObject):
    """
    Класс служит для восстановления расшота в новом монтаже.
    """
    finished = pyqtSignal(str) 
    progress = pyqtSignal(str)
    error = pyqtSignal(str) 

    def __init__(self, fps: int, project: str, edit_name: str, 
                 target_edit_path: str, logic: str):
        super().__init__()
        self.fps = fps
        self.project = project
        self.edit_name = edit_name
        self.target_edit = target_edit_path
        self.logic = logic

    def timecode_to_frame(self, fps: int, timecode: str) -> int:
        """
        Переводит таймкоды в значения фреймов.
        """
        return tc(fps, timecode).frames

    def frame_to_timecode(self, fps: int, frames: int) -> str:
        """
        Переводит фреймы в значения таймкодов.
        """
        return str(tc(fps, frames=frames))

    def overlap_range(self, fps: int, base_src_in: str, base_src_out: str, targ_src_in: str,
                       targ_src_out: str) -> bool:
        """
        Возвращает True, если два диапазона таймкодов пересекаются хотя бы в одном кадре.
        """
        base_in  = self.timecode_to_frame(fps, base_src_in)
        base_out = self.timecode_to_frame(fps, base_src_out)
        targ_in  = self.timecode_to_frame(fps, targ_src_in)
        targ_out = self.timecode_to_frame(fps, targ_src_out)

        return base_in <= targ_out and targ_in <= base_out
    
    def create_and_export_avid_loc(self, shot_info: tuple, output_path: Path, backup_output_path: Path) -> None:
        '''
        Создание и экспорт локатора AVID в аутпут файл.
        '''
        try:
            shot_data, shot_name = shot_info
            rec_in = self.timecode_to_frame(self.fps, shot_data.edl_record_in)
            rec_out = self.timecode_to_frame(self.fps, shot_data.edl_record_out)
            
            with open(output_path, "a", encoding='utf8') as o, open(backup_output_path, "a", encoding='utf8') as ob:
                timecode = int((rec_in + (rec_in + (rec_out - rec_in))) / 2)
                # Используется спец табуляция для корректного импорта в AVID
                output_string = f'PGM	{self.frame_to_timecode(self.fps, timecode)}	V3	yellow	{shot_name}\n'
                o.write(output_string)
                ob.write(output_string)
        except:
            message = f"Не удалось создать маркер для шота {shot_name}"
            logger.debug(message)
            self.progress.emit(message)

    def show_duplicates(self, processed_shots: dict) -> None:
        '''
        Метод находит выводит информацию в GUI об дубликатах шотов.
        '''
        self.progress.emit(f"Дубликаты шотов в EDL:\n")
        for shot_name, count in processed_shots.items():
            if len(count) >= 2:
                if shot_name in processed_shots:
                    for data in processed_shots[shot_name]:
                        self.progress.emit(f"Шот {shot_name}:   rec_in - {data.edl_record_in},  rec_out - {data.edl_record_out}")
                self.progress.emit("\n")

    def run(self) -> None:
        """
        Основная логика.
        """
        try:
            db = EditDatabase(DATA_PATH, self.project)

            if self.logic == "Edit":
                base_edit = db.get_shots_by_edit(self.project, self.edit_name)
            elif self.logic == "Actual":
                base_edit = db.get_shots_by_actual(self.project)

            target_edit = detect_edl_parser(self.fps, edl_path=self.target_edit)
            
            output_path = Path(str(self.target_edit).replace(".edl", f"_restored_rasshot_{d.today()}.edl"))
            backup_path = get_output_path(self.project, "edl", os.path.basename(self.target_edit).replace(".edl", f"_restored_rasshot"))

            loc_path = Path(str(self.target_edit).replace(".edl", f"_AVID_LOC_{d.today()}.txt"))
            loc_backup_path = get_output_path(self.project, "txt", os.path.basename(self.target_edit).replace(".edl", f"_AVID_LOC"))
            
            with open(loc_path, "w", encoding='utf8') as _, open(loc_backup_path, "w", encoding='utf8') as _:
                pass
            
            with open(output_path, "w", encoding="utf-8") as o, open(backup_path, "w", encoding="utf-8") as ob:
                processed_shots_tmp = {}
                for target_edit_data in target_edit:
                    for _, base_shot_data in base_edit.items():

                        if base_shot_data["src_name"] == target_edit_data.edl_source_name and self.overlap_range(
                            self.fps,
                            base_shot_data["src_in"], base_shot_data["src_out_full"],
                            target_edit_data.edl_source_in, target_edit_data.edl_source_out_src
                        ):  
                            self.create_and_export_avid_loc((target_edit_data, base_shot_data["shot_name"]), loc_path, loc_backup_path)

                            str1 = (f"{target_edit_data.edl_record_id} {base_shot_data['shot_name']} "
                                    f"{target_edit_data.edl_track_type} {target_edit_data.edl_transition} "
                                    f"{target_edit_data.edl_source_in} {target_edit_data.edl_source_out} "
                                    f"{target_edit_data.edl_record_in} {target_edit_data.edl_record_out}")
                            str2 = f'\n* FROM CLIP NAME: {base_shot_data["shot_name"]}\n'
                            o.write(str1)
                            ob.write(str1)
                            o.write(str2)
                            ob.write(str2)

                            processed_shots_tmp.setdefault(base_shot_data["shot_name"], []).append(target_edit_data)
                            break

            if processed_shots_tmp:                  
                self.show_duplicates(processed_shots_tmp)

            logger.info(f"Сформированы файлы: \n{output_path}\n{backup_path}\n{loc_path}\n{loc_backup_path}")
            self.finished.emit(f"Обработка завершена!")
        except Exception as e:
            self.error.emit(f"Ошибка: {e}")

class EDLInit(QObject):
    """
    Класс служит для инициализации данных из EDL.
    """
    finished = pyqtSignal(str) 
    error = pyqtSignal(str) 

    def __init__(self, fps: int, edl_path: str, project: str, update_status: bool,
                  settings_config: dict):
        super().__init__()
        self.fps = fps
        self.edl_path = edl_path
        self.project = project
        self.update_status = update_status
        self.settings_config = settings_config

    def run(self) -> None:
        """
        Основная логика.
        """
        try:
            db_path = DATA_PATH

        except Exception as e:
            self.error.emit(f"Ошибка получения пути к базе данных: {e}")
        
        try:
            db = EditDatabase(db_path, self.project)
        except Exception as e:
            self.error.emit(f"Ошибка получения объекта базы данных: {e}")

        try:
            parser_data = detect_edl_parser(self.fps, edl_path=self.edl_path)
        except Exception as e:
            self.error.emit(f"Ошибка парсинга EDL: {e}")

        try:
            for data in parser_data:
                match =  re.match(self.settings_config["patterns"]["compare_versions_shot_no_versions_mask"], data.edl_shot_name)
                if match:
                    db.add_shot(project=self.project,
                                shot_name=data.edl_shot_name,
                                edit_name=data.edl_edit_name,
                                shot_id=data.edl_record_id,
                                track_type=data.edl_track_type,
                                transition=data.edl_transition,
                                src_in=data.edl_source_in,
                                src_out=data.edl_source_out,
                                src_out_full=data.edl_source_out_src,
                                rec_in=data.edl_record_in,
                                rec_out=data.edl_record_out,
                                src_name=data.edl_source_name,
                                update_status=self.update_status
                                )

            db.save()
            db.backup()
            self.finished.emit("Данные успешно добавлены!")
        except Exception as e:
            self.error.emit(f"Ошибка добавления данных в базу: {e}")

class EDLComparator(QObject):
    """
    Класс служит для сравнения двух EDL и выдачи результата в виде отчета.
    """
    finished = pyqtSignal(str) 
    progress = pyqtSignal(str)
    error = pyqtSignal(str) 

    def __init__(self, fps: int, project: str, base_edit: str, 
                target_edit: str, base_logic: str, target_logic: str):
        super().__init__()
        self.fps = fps
        self.project = project
        self.base_edit_name = base_edit
        self.target_edit_name = target_edit
        self.base_logic = base_logic
        self.target_logic = target_logic

    def timecode_to_frame(self, fps: int, timecode: str) -> int:
        """
        Переводит таймкоды в значения фреймов.
        """
        return tc(fps, timecode).frames

    def frame_to_timecode(self, fps: int, frames: int) -> str:
        """
        Переводит фреймы в значения таймкодов.
        """
        return str(tc(fps, frames=frames))
    
    def out_hyper(self, file_path: str) -> None:
        """
        Выводит в GUI ссылку на аутпут документ.
        """
        url = Path(file_path).resolve().as_uri()
        self.progress.emit(f'Посмотреть отчет: <a href="{url}">{url}</a></span>')

    def overlap_range(self, fps: int, base_src_in: str, base_src_out: str, targ_src_in: str, 
                      targ_src_out: str, shot_name: str, target_shot_data: str) -> bool:
        """
        Сравнивает диапазоны таймкодов.
        Возвращает bool есть ли пересечение.
        """
        base_in = self.timecode_to_frame(fps, base_src_in)
        base_out = self.timecode_to_frame(fps, base_src_out)
        targ_in = self.timecode_to_frame(fps, targ_src_in)
        targ_out = self.timecode_to_frame(fps, targ_src_out)

        overlap_start = max(base_in, targ_in)
        overlap_end = min(base_out, targ_out)

        if overlap_start <= overlap_end:

            start_diff = targ_in - base_in   # Сдвиг начала
            end_diff = targ_out - base_out   # Сдвиг конца

            if start_diff == 0 and end_diff == 0:
                self.reedit_data.setdefault('No changes', []).append(shot_name)

            elif (
                (start_diff > 0 and end_diff < 0) or
                (start_diff > 0 and end_diff == 0) or
                (start_diff == 0 and end_diff < 0)
                    ):
                self.reedit_data.setdefault('Less', []).append((shot_name, ~start_diff + 1, end_diff))
                self.reedit_data_edl.setdefault('Less', []).append(target_shot_data)

            else:
                self.reedit_data.setdefault('More', []).append((shot_name, ~start_diff + 1, end_diff))
                self.reedit_data_edl.setdefault('More', []).append(target_shot_data)

            return True
        return False
    
    def is_leave(self) -> None:
        """
        Провеяем ушли ли шоты из нового монтажа.
        """
        for be_shot in self.base_edit:
            if be_shot not in self.target_edit.keys():
                self.reedit_data.setdefault('Leave', []).append(be_shot)

    def is_new(self) -> None:
        """
        Проверяем появились ли новые шоты в новом монтаже.
        """
        for te_shot in self.target_edit:
            if te_shot not in self.base_edit.keys():
                self.reedit_data.setdefault('New', []).append(te_shot)
                self.reedit_data_edl.setdefault('New', []).append(self.target_edit[te_shot])

    def export_to_excel(self) -> None:
        """
        Экспортирует reedit_data в Excel.
        Формат: Category | Shot | Start | End.
        """
        try:
            if not self.reedit_data:
                logger.warning("Нет данных для экспорта")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "re-edit report"

            # Заголовок
            header = ["Category", "Shot", "Start", "End"]
            ws.append(header)

            # Стили
            green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            red_fill   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            for category, items in self.reedit_data.items():
                for item in items:

                    #  More / Less
                    if category in ("More", "Less") and isinstance(item, tuple):
                        shot_name, start, end = item

                        row = [category, shot_name, start, end]
                        ws.append(row)

                        start_cell = ws[f"C{ws.max_row}"]
                        if start > 0:
                            start_cell.fill = green_fill
                        elif start < 0:
                            start_cell.fill = red_fill

                        end_cell = ws[f"D{ws.max_row}"]
                        if end > 0:
                            end_cell.fill = green_fill
                        elif end < 0:
                            end_cell.fill = red_fill

                    else:
                        row = [category, item, "-", "-"]
                        ws.append(row)

            # Автоширина
            for col in range(1, 5):
                letter = get_column_letter(col)
                max_len = max(len(str(cell.value)) for cell in ws[letter] if cell.value)
                ws.column_dimensions[letter].width = max_len + 2

            filepath = get_output_path(self.project, "xlsx", "reedit_report")

            wb.save(filepath)

            logger.info(f"Сохранен Excel файл: {filepath}")

            return filepath
        
        except Exception as e:
            raise 
        
    def create_output_edl(self, shot: dict, output) -> None:
        """
        Метод формирует аутпут файл в формате, пригодном для отображения оффлайн клипов в Resolve и AVID.

        :param output: Файловый объект.
        """
        str1 = (f"{shot['id']} {shot['shot_name']} "
                f"{shot['track_type']} {shot['transition']} "
                f"{shot['src_in']} {shot['src_out']} "
                f"{shot['rec_in']} {shot['rec_out']}")
        str2 = f"\n* FROM CLIP NAME: {shot['shot_name']}\n"
        output.write(str1)
        output.write(str2)

    def sort_output(self, reedit_data: dict) -> None:
        """
        Сортирует аутпут в зависимости от категории изменений.
        """
        for status, data in reedit_data.items():
            output_path = get_output_path(self.project, 'edl', status , subfolder='reedit_edl')
            with open(output_path, 'w', encoding="utf-8") as o:
                pass
            for shot_data in data:
                with open(output_path, 'a', encoding="utf-8") as o:
                    self.create_output_edl(shot_data, o)

            logger.info(f"Сформирован EDL файл: {output_path}")

    def find_cross(self) -> None:
        """
        Ищем пересекающиеся номера шотов и сравниваем значения.
        """
        for _, base_shot_data in self.base_edit.items():
            for _, target_shot_data in self.target_edit.items():
                # Сравниваем по именам шотов
                if base_shot_data["shot_name"] == target_shot_data["shot_name"]:
                    # Проверяем на предмет смены дубля
                    if base_shot_data["src_name"] == target_shot_data["src_name"]:
                        # Проверяем на предмет пересечения диапазонов
                        if self.overlap_range(
                            self.fps,
                            base_shot_data["src_in"], base_shot_data["src_out_full"],
                            target_shot_data["src_in"], target_shot_data["src_out_full"],
                            base_shot_data["shot_name"], target_shot_data
                        ):
                            break
                        else:
                            self.reedit_data_edl.setdefault('Phase changed', []).append(base_shot_data)
                    else:
                        self.reedit_data_edl.setdefault('Take changed', []).append(base_shot_data)

    def run(self) -> None:
        """
        Основная логика.
        """
        self.reedit_data = {}
        self.reedit_data_edl = {}
        try:
            db = EditDatabase(DATA_PATH, self.project)

            if self.base_logic == "Edit":
                self.base_edit = db.get_shots_by_edit(self.project, self.base_edit_name)
            if self.base_logic == "Actual":
                self.base_edit = db.get_shots_by_actual(self.project)
            
            if self.target_logic == "Edit":
                self.target_edit = db.get_shots_by_edit(self.project, self.target_edit_name)
            if self.target_logic == "Actual":
                self.target_edit = db.get_shots_by_actual(self.project)

            self.find_cross()

            self.is_leave()

            self.is_new()
            
            result_path = self.export_to_excel()

            self.sort_output(self.reedit_data_edl)
            
            self.out_hyper(result_path)

            logger.info(f"Сформирован отчет: {result_path}")
            self.finished.emit(f"Обработка успешно завершена!")
        except Exception as e:
            self.error.emit(f"Ошибка: {e}")

class PhaseChecker(QObject):
    """
    Класс служит для сравнения двух EDL и выдачи результата в виде отчета.
    """
    finished = pyqtSignal(str) 
    progress = pyqtSignal(str)
    error = pyqtSignal(str) 

    def __init__(self, fps: int, project: str, base_edit: str, 
                                    target_edits: str):
        super().__init__()
        self.fps = fps
        self.project = project
        self.base_edit = base_edit
        self.target_edits = target_edits

    def timecode_to_frame(self,timecode: str) -> int:
        """
        Переводит таймкоды в значения фреймов.
        """
        return tc(self.fps, timecode).frames

    def frame_to_timecode(self, frames: int) -> str:
        """
        Переводит фреймы в значения таймкодов.
        """
        return str(tc(self.fps, frames=frames))
    
    def create_edl(self, data) -> None:
        """
        Создание EDL файла.
        """
        for shot in data:
            id = shot["id"]
            src_name = shot["src_name"]
            track_type = shot["track_type"]
            transition = shot["transition"]
            src_in = shot["src_in"]
            src_out = shot["src_out"]
            rec_in = shot["rec_in"]
            rec_out = shot["rec_out"]

            with open("result_edl.edl", "a", encoding="utf-8") as o:
                o.write(f"{id} {src_name} "
                        f"{track_type} {transition} "
                        f"{src_in} {src_out} "
                        f"{rec_in} {rec_out}\n")

    def get_rec_out(self, src_in: str, src_out: str, rec_in: str) -> str:
        """
        Метод вычилсяет rec_out таймкод.
        """
        src_duration = self.timecode_to_frame(src_out) - self.timecode_to_frame(src_in)
        rec_out = self.frame_to_timecode(self.timecode_to_frame(rec_in) + src_duration)
        return rec_out

    def get_max_range(self, filtred_data: dict) -> list:
        """
        Метод ищет самый ранний таймкод source_in и самый поздный таймкод source_out, 
        высчитывает rec_in и rec_out и устанавливает полученные значения в донора.
        """
        result = []
        rec_in_data = []
        rec_in_default = "01:00:00:00"

        for shot_name, shot_data in filtred_data.items():
            # Берём копию донора, чтобы не затирать исходные данные
            edit_donor = dict(min(shot_data, key=lambda x: self.timecode_to_frame(x["src_in"])))

            # Выбираем минимальный src_in и максимальный src_out
            min_src_in_edit = min(shot_data, key=lambda x: self.timecode_to_frame(x["src_in"]))
            max_src_out_edit = max(shot_data, key=lambda x: self.timecode_to_frame(x["src_out_full"]))

            # Вычитаем 1 кадр из src_out для корректного отображения в интерфейсе программы
            src_out = self.frame_to_timecode(self.timecode_to_frame(max_src_out_edit['src_out_full']) - 1)
            self.progress.emit(f"Полный диапазон кадров шота {shot_name}: {min_src_in_edit['src_in']} - {src_out}")

            edit_donor["src_in"] = min_src_in_edit["src_in"]
            edit_donor["src_out_full"] = max_src_out_edit["src_out_full"]

            # Обновляем rec_in/rec_out
            if rec_in_data:
                edit_donor["rec_in"] = rec_in_data[-1]
            else:
                edit_donor["rec_in"] = rec_in_default

            edit_donor["rec_out"] = self.get_rec_out(
                edit_donor["src_in"],
                edit_donor["src_out_full"],
                edit_donor["rec_in"]
            )

            # Запоминаем конец текущего для следующего шота
            rec_in_data.append(edit_donor["rec_out"])

            result.append(edit_donor)

        return result

    def compare(self, base_data: dict, trg_data: dict) -> None:
        """
        Метод сравнивает сорс диапазон шота в базовом монтаже с диапазонами
        из выбранных для сравнения(trg_data_list) монтажей этого шота.
        Ищет только те случаи, когда в trg_data_list есть диапазоны шире чем в базовом.
        """
        base_src_in = self.timecode_to_frame(base_data["src_in"])
        base_src_out = self.timecode_to_frame(base_data["src_out_full"])
        shot_name = base_data["shot_name"]

        warnings = []
        for trg_data in trg_data:
            trg_src_in = self.timecode_to_frame(trg_data["src_in"])
            trg_src_out = self.timecode_to_frame(trg_data["src_out_full"])

            overlap_start = max(base_src_in, trg_src_in)
            overlap_end = min(base_src_out, trg_src_out)

            if overlap_start <= overlap_end:

                start_diff = trg_src_in - base_src_in   # Сдвиг начала
                end_diff = trg_src_out - base_src_out   # Сдвиг конца

                if start_diff < 0 or end_diff > 0:
                    self.filtred_data.setdefault(shot_name, []).append(trg_data)
                    # Добавляем базовый монтаж в случае срабатывания условия для сравнения
                    self.filtred_data.setdefault(shot_name, []).append(base_data)

        if warnings:
            no_dubl = str(*set(warnings))
            self.progress.emit(no_dubl)

    def run(self) -> None:
        """
        Основная логика.
        """
        try:
            db = EditDatabase(DATA_PATH, self.project)

            base_edit = db.get_shots_by_edit(self.project, self.base_edit)
            target_edits = db.get_shots_by_edits(self.project, self.target_edits)

            self.filtred_data = {}
            for base_shot_name, base_shot_data in base_edit.items():
                for trg_shot_name, trg_shot_data in target_edits.items():
                    if base_shot_name == trg_shot_name:
                        self.compare(base_shot_data, trg_shot_data)
                            
            if self.filtred_data:
                adjusted_ranges = self.get_max_range(self.filtred_data)
            
            #self.create_edl(adjusted_ranges)
            self.finished.emit(f"Обработка успешно завершена!")

        except Exception as e:
            self.error.emit(f"Ошибка: {e}")

class EDLGui(QWidget):

    def __init__(self):
        super().__init__()

        self.setWindowTitle("Edit Database")
        self.resize(800, 600)
        self.setWindowFlags(Qt.WindowStaysOnTopHint)


        # Главное меню с вкладками
        tabs = QTabWidget()
        tabs.addTab(self.init_tab(), "Init Edit")
        tabs.addTab(self.restore_tab(), "Restore shots")
        tabs.addTab(self.compare_tab(), "Compare Edits")
        tabs.addTab(self.check_phase_tab(), "Max Range")
        tabs.addTab(self.view_database(), "View Database")

        tabs.setStyleSheet("""
            QTabBar::tab {
                width: 100px;
            }
        """)

        layout = QVBoxLayout()
        layout.addWidget(tabs)
        self.setLayout(layout)

        self.thread = None
        self.worker = None

    def init_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignTop)

        # FPS input
        fps_layout = QHBoxLayout()
        fps_layout.addStretch()
        fps_layout.addWidget(QLabel("FPS:"))
        self.init_fps_input = QLineEdit("24")
        self.init_fps_input.setFixedWidth(40)
        fps_layout.addWidget(self.init_fps_input)
        fps_layout.addStretch()
        layout.addLayout(fps_layout)

        # Project name input
        project_layout = QHBoxLayout()
        project_layout.addStretch()
        self.project_combo = QComboBox()
        self.project_combo.addItems(self.get_project())
        self.project_combo.currentTextChanged.connect(self.get_project_settings)
        self.project_combo.setFixedWidth(300)
        project_layout.addWidget(self.project_combo)
        project_layout.addStretch()
        layout.addLayout(project_layout)

        # Create project
        create_project_layout = QHBoxLayout()
        create_project_layout.addStretch()
        self.project_edit_name = QLineEdit()
        self.project_edit_name.setPlaceholderText("Enter a name to create a project")
        self.project_edit_name.setFixedWidth(300)
        self.project_edit_name.returnPressed.connect(lambda: self.create_project())
        create_project_layout.addWidget(self.project_edit_name)
        create_project_layout.addStretch()
        layout.addLayout(create_project_layout)

        # EDL path input
        path_layout = QHBoxLayout()
        path_layout.addWidget(QLabel("EDL Path:"))
        path_layout.addSpacing(13)
        self.init_edl_input = QLineEdit()
        browse_button = QPushButton("Choose")
        browse_button.clicked.connect(self.browse_init_edl)
        path_layout.addWidget(self.init_edl_input)
        path_layout.addWidget(browse_button)
        layout.addLayout(path_layout)

        # Add button
        self.init_start_btn = QPushButton("Start")
        self.init_start_btn.clicked.connect(self.start_init)
        layout.addWidget(self.init_start_btn)

        tab.setLayout(layout)
        return tab

    def restore_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        # FPS input
        fps_layout = QHBoxLayout()
        fps_layout.addStretch()
        fps_layout.addWidget(QLabel("FPS:"))
        self.restore_fps_input = QLineEdit("24")
        self.restore_fps_input.setMaximumWidth(40)
        fps_layout.addWidget(self.restore_fps_input)
        fps_layout.addStretch()
        layout.addLayout(fps_layout)
        
        # Project
        restore_project_layout = QHBoxLayout()
        restore_project_layout.addStretch()
        self.restore_project_combo = QComboBox()
        self.restore_project_combo.setFixedWidth(300)
        self.restore_project_combo.addItems(self.get_project())
        restore_project_layout.addWidget(self.restore_project_combo)
        restore_project_layout.addStretch()
        layout.addLayout(restore_project_layout)

        # Base Edit
        base_layout = QHBoxLayout()
        base_layout.addWidget(QLabel("Base Edit:"))

        base_layout.addSpacing(10)
        self.logic_combo = QComboBox()
        self.logic_combo.addItems(["Based on", "Actual", "Edit"])
        base_layout.addWidget(self.logic_combo, 1)

        self.restore_edit_combo = QComboBox()
        self.restore_edit_combo.setMinimumWidth(250)
        base_layout.addWidget(self.restore_edit_combo, 1)

        self.restore_project_combo.currentTextChanged.connect(
            lambda _: self.get_edit(self.restore_project_combo, self.restore_edit_combo, self.logic_combo)
        )
        self.logic_combo.currentTextChanged.connect(
            lambda _: self.get_edit(self.restore_project_combo, self.restore_edit_combo, self.logic_combo)
        )

        layout.addLayout(base_layout)

        # Target Edit
        target_layout = QHBoxLayout()
        target_layout.addWidget(QLabel("Target Edit:"))
        base_layout.addSpacing(4)
        self.restore_new_input = QLineEdit()
        browse_button = QPushButton("Choose")
        browse_button.clicked.connect(self.browse_restore_new)
        target_layout.addWidget(self.restore_new_input)
        target_layout.addWidget(browse_button)
        layout.addLayout(target_layout)

        # Run button
        self.run_button = QPushButton("Start")
        self.run_button.clicked.connect(self.start_restore)
        layout.addWidget(self.run_button)

        # Log
        self.restore_log = QTextEdit()
        self.restore_log.setPlaceholderText("Restore warnings")
        self.restore_log.setReadOnly(True)
        layout.addWidget(self.restore_log)

        tab.setLayout(layout)
        return tab
    
    def compare_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        # FPS
        fps_layout = QHBoxLayout()
        fps_layout.addStretch()
        fps_layout.addWidget(QLabel("FPS:"))
        self.fps_input = QLineEdit("24")
        self.fps_input.setMaximumWidth(40)
        fps_layout.addWidget(self.fps_input)
        fps_layout.addStretch()
        layout.addLayout(fps_layout)

        # Project
        compare_project_layout = QHBoxLayout()
        compare_project_layout.addStretch()
        self.compare_project_cb = QComboBox()
        self.compare_project_cb.setFixedWidth(300)
        self.compare_project_cb.addItems(self.get_project())
        compare_project_layout.addWidget(self.compare_project_cb)
        compare_project_layout.addStretch()
        layout.addLayout(compare_project_layout)

        # Base Edit
        base_edit_layout = QHBoxLayout()
        base_edit_layout.addWidget(QLabel("Base Edit:"))

        base_edit_layout.addSpacing(10)
        self.base_logic_combo = QComboBox()
        self.base_logic_combo.addItems(["Based on", "Actual", "Edit"])
        base_edit_layout.addWidget(self.base_logic_combo, 1)

        self.compare_base_edit_cb = QComboBox()
        self.compare_base_edit_cb.setMinimumWidth(250)
        base_edit_layout.addWidget(self.compare_base_edit_cb, 1)

        self.compare_project_cb.currentTextChanged.connect(
            lambda _: self.get_edit(self.compare_project_cb, self.compare_base_edit_cb, self.base_logic_combo)
        )
        self.base_logic_combo.currentTextChanged.connect(
            lambda _: self.get_edit(self.compare_project_cb, self.compare_base_edit_cb, self.base_logic_combo)
        )

        layout.addLayout(base_edit_layout)

        # Target edit
        target_edit_layout = QHBoxLayout()
        target_edit_layout.addWidget(QLabel("Target Edit:"))

        self.target_logic_combo = QComboBox()
        self.target_logic_combo.addItems(["Based on", "Actual", "Edit"])
        target_edit_layout.addWidget(self.target_logic_combo, 1)

        self.compare_target_edit_cb = QComboBox()
        self.compare_target_edit_cb.setMinimumWidth(250)
        target_edit_layout.addWidget(self.compare_target_edit_cb, 1)

        self.compare_project_cb.currentTextChanged.connect(
            lambda _: self.get_edit(self.compare_project_cb, self.compare_target_edit_cb, self.target_logic_combo)
        )
        self.target_logic_combo.currentTextChanged.connect(
            lambda _: self.get_edit(self.compare_project_cb, self.compare_target_edit_cb, self.target_logic_combo)
        )

        layout.addLayout(target_edit_layout)

        # Run button
        self.compare_start_btn = QPushButton("Start")
        self.compare_start_btn.clicked.connect(self.start_comparison)
        layout.addWidget(self.compare_start_btn)

        # Log
        self.log = QTextBrowser()
        self.log.setPlaceholderText("Compare report")
        self.log.setReadOnly(True)
        self.log.setOpenLinks(False)
        self.log.anchorClicked.connect(self.open_in_file_manager)
        layout.addWidget(self.log)

        tab.setLayout(layout)
        return tab
    
    def check_phase_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignTop)

        # FPS input
        fps_layout = QHBoxLayout()
        fps_layout.addStretch()
        fps_layout.addWidget(QLabel("FPS:"))
        self.check_fps_input = QLineEdit("24")
        self.check_fps_input.setMaximumWidth(40)
        fps_layout.addWidget(self.check_fps_input)
        fps_layout.addStretch()
        layout.addLayout(fps_layout)

        # Project
        check_project_layout = QHBoxLayout()
        check_project_layout.addStretch()
        self.check_project_cb = QComboBox()
        self.check_project_cb.setFixedWidth(300)
        self.check_project_cb.addItems(self.get_project())
        check_project_layout.addWidget(self.check_project_cb)
        check_project_layout.addStretch()
        layout.addLayout(check_project_layout)

        # Base Edit selection
        base_edit_layout = QHBoxLayout()
        base_edit_layout.addWidget(QLabel("Base Edit:"))
        base_edit_layout.addSpacing(10)
        self.check_base_edit_cb = QComboBox()

        base_edit_layout.addWidget(self.check_base_edit_cb, 1)

        self.check_project_cb.currentTextChanged.connect(
            lambda x: self.get_edit(
                self.check_project_cb,
                self.check_base_edit_cb,
                None
            )
        )

        layout.addLayout(base_edit_layout)

        # --- Target edits dynamic block ---
        self.target_edits_layout = QVBoxLayout()
        self.target_edits_rows = []

        # Первая строка создается сразу
        self.add_target_edit_row()

        # Кнопки Add / Remove
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()
        self.add_target_btn = QPushButton("Add Edit")
        self.add_target_btn.clicked.connect(self.add_target_edit_row)
        self.remove_target_btn = QPushButton("Remove Edit")
        self.remove_target_btn.clicked.connect(self.remove_target_edit_row)
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.add_target_btn)
        buttons_layout.addWidget(self.remove_target_btn)

        # Оборачиваем, чтобы кнопки всегда были внизу блока
        self.target_edits_layout.addLayout(buttons_layout)

        layout.addLayout(self.target_edits_layout)

        # Run button
        self.check_start_btn = QPushButton("Start")
        self.check_start_btn.clicked.connect(self.start_check_phase)
        layout.addWidget(self.check_start_btn)

        # Log output
        self.check_log = QTextEdit()
        self.check_log.setPlaceholderText("Check report")
        self.check_log.setReadOnly(True)
        layout.addWidget(self.check_log)

        self.get_edit(self.check_project_cb, self.check_base_edit_cb, None)
        tab.setLayout(layout)
        return tab
    
    def add_target_edit_row(self):
        """
        Добавить строку для выбора таргет-монтажа.
        """
        row_layout = QHBoxLayout()

        label = QLabel("Target Edit:")
        row_layout.addWidget(label)

        edit_cb = QComboBox()
        row_layout.addWidget(edit_cb, 1)

        self.check_project_cb.currentTextChanged.connect(
            lambda x, p=self.check_project_cb, e=edit_cb, l=None: self.get_edit(p, e, l)
        )

        # Вставляем перед кнопками Add/Remove
        self.target_edits_layout.insertLayout(len(self.target_edits_layout) - 1, row_layout)

        # Сохраняем всё вместе для корректного удаления
        self.target_edits_rows.append((row_layout, label, edit_cb))

        self.get_edit(self.check_project_cb, edit_cb, None)

    def remove_target_edit_row(self):
        """
        Удалить последнюю строку target edits (вместе с лейблом).
        """
        if self.target_edits_rows:
            row_layout, label, edit_cb = self.target_edits_rows.pop()

            # Удаляем все виджеты из layout
            for widget in (label, edit_cb):
                row_layout.removeWidget(widget)
                widget.deleteLater()

            # Удаляем сам layout из родительского
            self._delete_layout(row_layout)

    def _delete_layout(self, layout):
        """Полностью удалить QLayout и его содержимое."""
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
            elif item.layout() is not None:
                self._delete_layout(item.layout())
        layout.deleteLater()

    def view_database(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        # --- Search row ---
        search_layout = QHBoxLayout()

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Shot number")
        self.search_input.setFixedWidth(150)

        self.btn_prev = QPushButton("Prev")
        self.btn_prev.setFixedWidth(55)

        self.btn_next = QPushButton("Next")
        self.btn_next.setFixedWidth(55)

        self.btn_prev.clicked.connect(lambda: self.navigate_found(-1))
        self.btn_next.clicked.connect(lambda: self.navigate_found(1))

        self.btn_prev.setEnabled(False)
        self.btn_next.setEnabled(False)

        self.search_input.returnPressed.connect(self.search_shots)

        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.btn_prev)
        search_layout.addWidget(self.btn_next)
        search_layout.addStretch()

        layout.addLayout(search_layout)

        # --- End search row ---

        self.tree = QTreeView()
        layout.addWidget(self.tree)

        self.btn_save = QPushButton("Save changes")
        self.btn_save.clicked.connect(lambda: self.btn_save.setEnabled(False))
        self.btn_save.clicked.connect(self.save_json)
        layout.addWidget(self.btn_save)

        self.data = None

        if not os.path.exists(DATA_PATH):
            return tab

        self.load_json_from_path(DATA_PATH)

        return tab
    
    def open_in_file_manager(self, url:QUrl):
        """
        Метод открывает файл в файловом менеджере.
        """
        try:
            path = Path(url.toLocalFile())

            if sys.platform == 'win32': 
                os.startfile(path)
            else: 
                subprocess.Popen(['open', path])
        except Exception as e:
            self.on_error(self, "Error", f"Ошибка при открытии файла: {e}")
    
    def get_project_settings(self):
        """
        Получаем проектный конфиг
        """
        project_name = self.project_combo.currentText()
        load_config(project_name)
        self.config = get_config()

    def set_row_height(self, height: int):
        """
        Локальный стиль для вкладки View Database.
        """
        self.tree.setUniformRowHeights(True)
        self.tree.setStyleSheet(f"QTreeView::item {{ height: {height}px; }}")
        self.tree.setIconSize(QSize(height - 4, height - 4))

    def load_json_from_path(self, database_path):
        """
        Загружает датабазу во вкладке View Database.
        """
        self.database_path = database_path

        with open(database_path, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if not content:
                self.data = {} 
            else:
                self.data = json.loads(content)
                
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(["Project / Shot / Edit", "Shot data"])

        self.build_tree(self.model.invisibleRootItem(), self.data)

        self.tree.setModel(self.model)
        self.set_row_height(30)
        self.tree.header().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tree.setModel(self.model)
        self.tree_model = self.model 

    def build_tree(self, parent, data):
        """
        Метод строит таблицу дерева датабазы во вкладке View Database.
        """
        if isinstance(data, dict):
            for key, value in data.items():
                key_item = QStandardItem(str(key))
                key_item.setEditable(False)
                if isinstance(value, (dict, list)):
                    value_item = QStandardItem("")
                    self.build_tree(key_item, value)
                else:
                    value_item = QStandardItem(str(value))
                    value_item.setEditable(True)
                parent.appendRow([key_item, value_item])

        elif isinstance(data, list):
            for index, value in enumerate(data):
                key_item = QStandardItem(f"[{index}]")
                key_item.setEditable(False)
                if isinstance(value, (dict, list)):
                    value_item = QStandardItem("")
                    self.build_tree(key_item, value)
                else:
                    value_item = QStandardItem(str(value))
                    value_item.setEditable(True)
                parent.appendRow([key_item, value_item])

    def save_json(self):
        """
        Сохраняет изменения во вкладке View Database.
        """
        try:
            def read_tree(item):
                data = {}
                for row in range(item.rowCount()):
                    key = item.child(row, 0).text()
                    val_item = item.child(row, 1)

                    if item.child(row, 0).hasChildren():
                        data[key] = read_tree(item.child(row, 0))
                    else:
                        text = val_item.text()
                        if text.lower() == "false":
                            data[key] = False
                        elif text.lower() == "true":
                            data[key] = True
                        else:
                            data[key] = text

                return data

            root = self.model.invisibleRootItem()
            new_data = read_tree(root)

            with open(self.database_path, "w", encoding="utf-8") as f:
                json.dump(new_data, f, ensure_ascii=False, indent=4)

            self.on_finished(f"Изменения успешно сохранены: {self.database_path}")
            self.btn_save.setEnabled(True)
        except:
            with open(self.database_path, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=4)

            self.on_error(f"Не удалось сохранить изменения")
            self.btn_save.setEnabled(True)

    def search_shots(self):
        query = self.search_input.text().strip()
        if not query:
            return

        self.found_indexes = []
        root = self.tree_model.invisibleRootItem()

        def recursive_search(item):
            for row in range(item.rowCount()):
                child = item.child(row, 0)
                if query in child.text():
                    self.found_indexes.append(self.tree_model.indexFromItem(child))

                if child.hasChildren():
                    recursive_search(child)

        recursive_search(root)

        if not self.found_indexes:
            self.on_error(f"Шотов с номером '{query}' не найдено")
            return

        self.current_found = 0
        self.select_found_item(0)

        # Активируем навигацию при множественных совпадениях
        self.btn_prev.setEnabled(len(self.found_indexes) > 1)
        self.btn_next.setEnabled(len(self.found_indexes) > 1)

    def select_found_item(self, index_pos):
        if not self.found_indexes:
            return

        index = self.found_indexes[index_pos]

        # Развернуть родителей
        parent = index.parent()
        while parent.isValid():
            self.tree.expand(parent)
            parent = parent.parent()

        # Выделить и проскроллить
        self.tree.setCurrentIndex(index)
        self.tree.scrollTo(index)

    def navigate_found(self, direction):
        """
        direction = -1 (Prev), direction = +1 (Next)
        """
        if not hasattr(self, "found_indexes") or not self.found_indexes:
            return

        self.current_found += direction

        # зацикливание
        if self.current_found < 0:
            self.current_found = len(self.found_indexes) - 1
        elif self.current_found >= len(self.found_indexes):
            self.current_found = 0

        self.select_found_item(self.current_found)

    def browse_restore_new(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select New Edit", "", "EDL files (*.edl)")
        if path:
            self.restore_new_input.setText(path)
    
    def browse_init_edl(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select EDL", "", "EDL files (*.edl)")
        if path:
            self.init_edl_input.setText(path)

    def add_to_data(self):
        edl_path = self.init_edl_input.text().strip()
        if not os.path.isfile(edl_path):
            QMessageBox.critical(self, "Error", "Укажите корректный EDL файл")
            return
        # Здесь можно вставить логику добавления в базу
        QMessageBox.information(self, "Success", f"EDL добавлен: {edl_path}")

    # --- Методы для вкладки Analyse ---
    def browse_prev(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Previous Edit", "", "EDL files (*.edl)")
        if path:
            self.prev_input.setText(path)

    def browse_new(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select New Edit", "", "EDL files (*.edl)")
        if path:
            self.new_input.setText(path)

    def browse_out(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save as", "", "EDL files (*.edl)")
        if path:
            self.out_input.setText(path)

    def get_edit(self, project_cb, edit_cb, logic_combo):
        if logic_combo is None or logic_combo.currentText() == "Edit":
            edit_cb.setEnabled(True)

            project_name = project_cb.currentText()
            if not project_name or project_name == "Select Project":
                edit_cb.clear()
                edit_cb.addItem("Select Edit")
                return

            db = EditDatabase(DATA_PATH, project_name)

            edits_list = db.get_edits(project_name) or []
            edits_list.insert(0, "Select Edit")

            edit_cb.clear()
            edit_cb.addItems(edits_list)
        else:
            edit_cb.setEnabled(False)
            edit_cb.clear()

    def get_project(self):
        """
        Метод получает список проектов из корневого каталога.
        """
        project_path = {"win32": GLOBAL_CONFIG["paths"]["root_projects_win"], 
            "darwin": GLOBAL_CONFIG["paths"]["root_projects_mac"]}[sys.platform] 
        if os.path.exists(project_path):
            projects_list = sorted([i for i in os.listdir(Path(project_path)) if os.path.isdir(Path(project_path) / i)])
            projects_list.insert(0, "Select Project")
            return projects_list
        else:
            self.on_error("Путь к папке с проекта не обнаружен")
            return

    def init_validate_inputs(self):
        """
        Валидация пользовательских данных из таба Init Edit.
        """
        try:
            fps = int(self.init_fps_input.text())
            if fps <= 0:
                raise ValueError("FPS должно быть больше 0")
        except ValueError:
            self.on_error("FPS должно быть целым числом больше 0")
            return None

        edl_path = self.init_edl_input.text().strip()
        if not os.path.isfile(edl_path):
            self.on_error("Укажите корректный путь к EDL")
            return None

        return fps, edl_path

    def start_init(self):
        """
        Запуск процесса добавления EDL в базу.
        """
        inputs = self.init_validate_inputs()
        if not inputs:
            return

        fps, edl_path = inputs
        project = self.project_combo.currentText()

        answer = QMessageBox.question(
            self,
            "Добавление монтажа",
            "Присвоить этому монтажу статус 'Actual'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )

        update_status = (answer == QMessageBox.Yes)

        self.thread = QThread()
        self.worker = EDLInit(fps, edl_path, project, update_status, self.config) 
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.init_start_btn.setEnabled(False)
        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)

        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.finished.connect(lambda: self.init_start_btn.setEnabled(True))
        self.thread.finished.connect(lambda: self.load_json_from_path(DATA_PATH))
        self.worker.error.connect(lambda: self.init_start_btn.setEnabled(True))
        self.worker.error.connect(lambda: self.thread.quit())

        self.thread.start()

    def restore_validate_inputs(self):
        """
        Валидация пользовательских данных из таба Restore Shots.
        """
        try:
            fps = int(self.restore_fps_input.text())
            if fps <= 0:
                raise ValueError
        except ValueError:
            self.on_error("FPS должно быть целым числом больше 0")
            return None
        
        logic = self.logic_combo.currentText().strip()
        if logic == "Based on":
            self.on_error("Выберете корректное значение логики")
            return None
        
        project = self.restore_project_combo.currentText()
        if project == "Select Project":
            self.on_error("Укажите проект для базового монтажа")
            return None

        if self.logic_combo.currentText() == "Edit":
            edit_name = self.restore_edit_combo.currentText().strip()
            if edit_name == "Select Edit":
                self.on_error("Укажите базовый монтаж")
                return None
        else:
            edit_name = None

        target_edit_path = self.restore_new_input.text().strip()
        if not os.path.isfile(target_edit_path):
            self.on_error("Укажите корректный путь к Target Edit")
            return None
               
        return fps, project, edit_name, target_edit_path, logic

    def start_restore(self):
        """
        Запуск восстановления шотов в новом монтаже.
        """
        inputs = self.restore_validate_inputs()
        if not inputs:
            return

        fps, project, edit_name, target_edit_path, logic = inputs

        self.thread = QThread()
        self.worker = ShotRestorer(fps, project, edit_name, target_edit_path, logic)
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.run_button.setEnabled(False)
        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)
        self.worker.progress.connect(self.restore_log.append)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.finished.connect(lambda: self.run_button.setEnabled(True))
        self.worker.error.connect(lambda: self.run_button.setEnabled(True))
        self.worker.error.connect(lambda: self.thread.quit())

        self.thread.start()

    def analyze_validate_inputs(self):
        """
        Валидация пользовательских данных из таба Compare Edits.
        """
        project = self.compare_project_cb.currentText().strip()
        base_edit = self.compare_base_edit_cb.currentText().strip()
        target_edit = self.compare_target_edit_cb.currentText().strip()
        base_logic = self.base_logic_combo.currentText().strip()
        target_logic = self.target_logic_combo.currentText().strip()

        try:
            fps = int(self.fps_input.text())
            if fps <= 0:
                raise ValueError("FPS должно быть больше 0")
        except ValueError:
            self.on_error("FPS должно быть целым числом больше 0")
            return None

        if base_logic == "Based on":
            self.on_error("Выберете корректное значение логки")
            return None

        if project == "Select Project":
            self.on_error("Укажите проект для базового монтажа")
            return None

        if base_logic == "Edit":
            if base_edit == "Select Edit":
                self.on_error("Укажите базовый монтаж")
                return None
            
        if target_logic == "Based on":
            self.on_error("Выберете корректное значение логики")
            return None
        
        if target_logic == "Edit":
            if target_edit == "Select Edit":
                self.on_error("Укажите целевой монтаж")
                return None
    
        return fps, project, base_edit, target_edit, base_logic, target_logic

    def start_comparison(self):
        """
        Запуск сравнения монтажей.
        """
        inputs = self.analyze_validate_inputs()
        if not inputs:
            return

        fps, project, base_edit, target_edit, base_logic, target_logic = inputs
        self.thread = QThread()
        self.worker = EDLComparator(fps, project, base_edit, 
                                    target_edit, base_logic, target_logic)
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.compare_start_btn.setEnabled(False)
        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)
        self.worker.progress.connect(self.log.append)

        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.finished.connect(lambda: self.compare_start_btn.setEnabled(True))
        self.worker.error.connect(lambda: self.compare_start_btn.setEnabled(True))
        self.worker.error.connect(lambda: self.thread.quit())

        self.thread.start()

    def check_validate_inputs(self):
        """
        Валидация пользовательских данных из вкладки Check Phase.
        """
        # FPS
        try:
            fps = int(self.check_fps_input.text())
            if fps <= 0:
                raise ValueError
        except ValueError:
            self.on_error("FPS должно быть целым числом больше 0")
            return None


        project = self.check_project_cb.currentText().strip()
        if project == "Select Project":
            self.on_error("Укажите проект")
            return None

        base_edit = None
        base_edit = self.check_base_edit_cb.currentText().strip()
        if base_edit == "Select Edit":
            self.on_error("Укажите Base Edit")
            return None

        target_edits = []
        for idx, (layout, label, edit_cb) in enumerate(self.target_edits_rows, start=1):
            edit_name = None
            edit_name = edit_cb.currentText().strip()
            if edit_name == "Select Edit":
                self.on_error(f"Укажите монтаж для Target Edit #{idx}")
                return None

            target_edits.append((edit_name))

        if not target_edits:
            self.on_error("Добавьте хотя бы один Target Edit")
            return None

        return fps, project, base_edit, target_edits

    def start_check_phase(self):
        """
        Запуск проверки фаз (логика для вкладки Check Phase).
        """
        inputs = self.check_validate_inputs()
        if not inputs:
            return
        
        target_edits = []
        for row_layout, label, edit_cb in self.target_edits_rows:
            target_edits.append((
                edit_cb.currentText()
            ))

        fps, project, base_edit, target_edits = inputs

        self.thread = QThread()
        self.worker = PhaseChecker(fps, project, base_edit, target_edits)
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.check_start_btn.setEnabled(False)
        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)
        self.worker.progress.connect(self.check_log.append)

        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.finished.connect(lambda: self.check_start_btn.setEnabled(True))
        self.worker.error.connect(lambda: self.check_start_btn.setEnabled(True))
        self.worker.error.connect(lambda : self.thread.quit())

        self.thread.start()

    def create_project(self):
        """
        Создает новый фолдер(проект)
        """
        new_project = self.project_edit_name.text().lower()
        base_path = Path({"win32": GLOBAL_CONFIG["paths"]["root_projects_win"], 
                        "darwin": GLOBAL_CONFIG["paths"]["root_projects_mac"]}[sys.platform])
        new_project_path = base_path / new_project
        new_project_path.mkdir(exist_ok=True)

        self.project_combo.clear()
        self.project_combo.addItems(self.get_project())

    def on_finished(self, message: str):
        QMessageBox.information(self, "Success", message)
        logger.info(message)

    def on_error(self, message):
        QMessageBox.critical(self, "Error", message)
        logger.info(message)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    apply_style(app)
    window = EDLGui()
    window.show()
    sys.exit(app.exec_())
